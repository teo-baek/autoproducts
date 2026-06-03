"""ezmerce v2 백엔드 워크스루 노트북 생성기 (nbformat).
실행: cd backend && .venv/bin/python notebooks/_build_notebook.py
"""
import nbformat as nbf

nb = nbf.v4.new_notebook()
cells = []


def md(src):
    cells.append(nbf.v4.new_markdown_cell(src))


def code(src):
    cells.append(nbf.v4.new_code_cell(src.strip("\n")))


md("""# ezmerce v2 백엔드 — 테스트 & 라이브 워크스루

이 노트북은 1차 백엔드의 **각 기능을 실제로 실행**해 보고, 동시에 **해당 pytest를 직접 돌려** 결과를 보여줍니다.

- 커널: **ezmerce v2 (.venv)** 를 선택하세요 (backend/.venv).
- 위에서부터 순서대로 실행하면 됩니다. 셀마다 `라이브 데모` + `pytest 실행`이 함께 들어 있습니다.

| 영역 | 데모 | 테스트 파일 |
|---|---|---|
| 품번 정규화 | platform_code 포맷 | test_platform_code.py |
| 가격 차등 노출 | 역할×유형 매트릭스 | test_pricing.py |
| 엑셀 파서 | 샘플 xlsx 파싱 | test_excel_parse.py |
| 이미지 매칭 | 파일명→품번 | test_image_match.py |
| QR | 이미지 직접 렌더 | test_qr.py |
| 엑셀 출력 | QR 삽입 xlsx | test_excel_export.py |
| 인증/RBAC | 권한 가드 | test_auth_dep.py / test_rbac.py |
| 카탈로그 | 역할별 셰이핑 | test_catalog_shaping.py |
| FastAPI 앱 | /health + 라우트 | test_health.py |
""")

md("## 0. 환경 셋업\n`backend/`를 작업 경로로 잡고 `app` 패키지를 import 가능하게 합니다.")
code("""
from pathlib import Path
import os, sys

here = Path.cwd()
BACKEND = here if (here / "app").exists() else (here.parent if (here.parent / "app").exists() else here)
os.chdir(BACKEND)
sys.path.insert(0, str(BACKEND))
print("backend :", BACKEND)
print("python  :", sys.version.split()[0])

import subprocess

def run_pytest(*files, q=False):
    \"\"\"해당 테스트 파일을 실제로 실행하고 출력을 보여준다.\"\"\"
    cmd = [sys.executable, "-m", "pytest", *files, "-v" if not q else "-q", "--no-header"]
    r = subprocess.run(cmd, capture_output=True, text=True)
    print(r.stdout[-3500:])
    if r.returncode != 0:
        print("STDERR:", r.stderr[-1500:])
    return r.returncode
""")

md("## 1. 전체 테스트 한 번에 (26 cases)\n외부 Supabase 없이 fake repo로 전부 돌아갑니다.")
code("""
run_pytest(q=True)
""")

md("## 2. 품번 정규화 — `platform_code`\n여러 도매업체 상품이 섞여도 충돌하지 않도록 플랫폼이 글로벌 식별자를 발급합니다.")
code("""
from app.services.platform_code import format_platform_code

for seq in (1, 42, 123456, 12345678):
    print(format_platform_code(seq, prefix="EZM"))

print("\\n--- pytest ---")
run_pytest("tests/test_platform_code.py")
""")

md("## 3. 가격 차등 노출 렌즈 (FR-5.2)\n같은 SKU라도 **보는 역할에 따라 다른 가격**이 나갑니다. 서버에서 결정(클라이언트 신뢰 X).")
code("""
import pandas as pd
from app.services.pricing import visible_price

SKU = {"wholesale_price": 12000, "retail_price": 29000, "product_org": "org-1"}
cases = [
    ("retail_seller", "independent", "org-1"),
    ("retail_seller", "agency_affiliated", "org-1"),
    ("agency", None, None),
    ("wholesaler", None, "org-1"),
    ("admin", None, None),
    ("guest", None, None),
]
rows = [{"role": r, "seller_type": st, "노출 결과": visible_price(r, st, SKU, viewer_org=o)} for r, st, o in cases]
display(pd.DataFrame(rows))

print("--- pytest ---")
run_pytest("tests/test_pricing.py")
""")

md("## 4. 표준 엑셀 템플릿 파서 (FR-2.2)\n표준 양식을 파싱하고, 도매가 누락 같은 불량 행은 에러로 분리해 수작업 매칭으로 넘깁니다.")
code("""
import openpyxl, tempfile, os, pandas as pd
from app.services.excel_parse import parse_template_rows, TEMPLATE_COLUMNS

wb = openpyxl.Workbook(); ws = wb.active
ws.append(TEMPLATE_COLUMNS)
ws.append(["1001", "린넨 셔츠", "화이트", "F", "12000", "29000"])
ws.append(["1002", "와이드 팬츠", "블랙", "L", "", "20000"])   # 도매가 누락 → error
p = os.path.join(tempfile.mkdtemp(), "sample.xlsx"); wb.save(p)

res = parse_template_rows(p)
print("정상 행:"); display(pd.DataFrame(res.rows))
print("에러 행:"); display(pd.DataFrame(res.errors))

print("--- pytest ---")
run_pytest("tests/test_excel_parse.py")
""")

md("## 5. 이미지 자동 매칭 (FR-2.3)\n업로드 이미지 파일명에서 품번 토큰을 뽑아 상품과 자동 연결합니다.")
code("""
from app.services.image_match import match_filename_to_product

products = {"1001": "product-A", "1002": "product-B"}
for fn in ["1001.jpg", "1001_main.png", "1002-01.jpeg", "zzz.png"]:
    print(f"{fn:16} -> {match_filename_to_product(fn, products)}")

print("\\n--- pytest ---")
run_pytest("tests/test_image_match.py")
""")

md("## 6. QR 코드 생성 (FR-4) — 이미지 직접 보기\nQR이 가리키는 공개 카드 URL과 실제 QR PNG를 인라인으로 렌더링합니다.")
code("""
from app.services.qr import qr_target_url, generate_qr_png
from IPython.display import Image, display

url = qr_target_url("EZM-000001", "https://shop.ezmerce.io")
print("QR 대상 URL:", url)
png = generate_qr_png(url)
print("PNG bytes:", len(png), "| 시그니처 OK:", png[:8] == b"\\x89PNG\\r\\n\\x1a\\n")
display(Image(data=png))

print("--- pytest ---")
run_pytest("tests/test_qr.py")
""")

md("## 7. 엑셀 출력 + QR 삽입 (FR-3)\n셀러용 카탈로그 엑셀의 **최우측 열에 QR 이미지가 자동 삽입**됩니다 (폼텍 라벨 대체).")
code("""
import openpyxl, tempfile, os
from app.services.excel_export import build_catalog_xlsx

items = [
    {"platform_code": "EZM-000001", "item_name": "린넨 셔츠", "price": 12000},
    {"platform_code": "EZM-000002", "item_name": "와이드 팬츠", "price": 18000},
]
out = os.path.join(tempfile.mkdtemp(), "catalog.xlsx")
build_catalog_xlsx(items, out, base_url="https://shop.ezmerce.io")

wb = openpyxl.load_workbook(out); ws = wb.active
print("헤더:", [c.value for c in ws[1]], "  (최우측 =", [c.value for c in ws[1]][-1], ")")
print("삽입된 QR 이미지 개수:", len(ws._images))
print("저장 위치:", out)

print("--- pytest ---")
run_pytest("tests/test_excel_export.py")
""")

md("## 8. 인증 / RBAC (FR-1)\n역할/승인 가드가 허용·차단을 올바로 수행하는지 직접 호출해 봅니다.")
code("""
from app.core.rbac import require_role, require_approved
from app.schemas.auth import CurrentUser

u = CurrentUser(id="u1", role="wholesaler", status="approved")
print("wholesaler -> wholesaler 권한:", require_role("wholesaler")(u).role, "(통과)")

try:
    require_role("admin")(u)
except Exception as e:
    print("wholesaler -> admin 시도:", type(e).__name__, getattr(e, "status_code", ""))

try:
    require_approved(CurrentUser(id="u2", role="agency", status="pending"))
except Exception as e:
    print("pending 계정 접근 시도:", type(e).__name__, getattr(e, "status_code", ""))

print("\\n--- pytest ---")
run_pytest("tests/test_auth_dep.py", "tests/test_rbac.py")
""")

md("## 9. 폐쇄형 카탈로그 셰이핑 (FR-5)\n같은 카탈로그 행이 역할에 따라 어떻게 다르게 보이는지 확인합니다.")
code("""
from app.routers.catalog import shape_catalog_item
from app.schemas.auth import CurrentUser

ROW = {"platform_code": "EZM-1", "item_name": "셔츠",
       "skus": [{"color": "화이트", "size": "F", "wholesale_price": 12000, "retail_price": 29000, "product_org": "org-9"}]}

for role, st in [("retail_seller", "independent"), ("retail_seller", "agency_affiliated"), ("agency", None)]:
    u = CurrentUser(id="u", role=role, status="approved", seller_type=st)
    item = shape_catalog_item(ROW, u)
    print(f"{role:14}/{str(st):18} -> {item['skus'][0]}")

print("\\n--- pytest ---")
run_pytest("tests/test_catalog_shaping.py")
""")

md("## 10. FastAPI 앱 — `/health` + 라우트 테이블\n앱을 띄워 헬스체크하고 등록된 라우트를 확인합니다.")
code("""
from fastapi.testclient import TestClient
from app.main import app

c = TestClient(app)
print("/health ->", c.get("/health").json())
print("\\n등록된 라우트:")
for path in sorted({r.path for r in app.routes}):
    print("  ", path)

print("\\n--- pytest ---")
run_pytest("tests/test_health.py")
""")

md("""---
✅ 여기까지 모두 통과하면 1차 백엔드 코어가 정상 동작하는 것입니다.

**아직 라우트가 없는 부분(로직은 구현됨):** 업로드 오케스트레이션(`/uploads/*`), `POST /auth/register`, `GET /catalog/export.xlsx`.
**런타임 연결:** Supabase에 `backend/migrations/2026-06-03_v2_core.sql` 적용 + `backend/.env`(SUPABASE_URL / SERVICE_KEY / JWT_SECRET).
""")

nb["cells"] = cells
nb["metadata"]["kernelspec"] = {
    "display_name": "ezmerce v2 (.venv)",
    "language": "python",
    "name": "ezmerce-v2",
}
nb["metadata"]["language_info"] = {"name": "python"}

out_path = "notebooks/ezmerce_v2_walkthrough.ipynb"
with open(out_path, "w", encoding="utf-8") as f:
    nbf.write(nb, f)
print("WROTE", out_path, "with", len(cells), "cells")
