import io
import logging
import time
from collections import OrderedDict

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from app.services.image_process import process_image_bytes
from app.services.qr import qr_target_url, generate_qr_png

log = logging.getLogger(__name__)

HEADERS = ["품번", "상품명", "가격", "QR"]


def _build_workbook(items: list[dict], base_url: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(HEADERS)
    for i, it in enumerate(items, start=2):
        ws.cell(row=i, column=1, value=it["platform_code"])
        ws.cell(row=i, column=2, value=it["item_name"])
        ws.cell(row=i, column=3, value=it.get("price"))
        png = generate_qr_png(qr_target_url(it["platform_code"], base_url))
        img = XLImage(io.BytesIO(png)); img.width = img.height = 64
        ws.add_image(img, f"D{i}")        # 최우측 열(QR)에 삽입
        ws.row_dimensions[i].height = 50
    return wb


def build_catalog_xlsx(items: list[dict], out_path: str, base_url: str) -> str:
    _build_workbook(items, base_url).save(out_path)
    return out_path


def catalog_xlsx_bytes(items: list[dict], base_url: str) -> bytes:
    """파일 경로 없이 메모리(BytesIO)로 xlsx 바이트 생성 — HTTP 다운로드 응답용."""
    buf = io.BytesIO()
    _build_workbook(items, base_url).save(buf)
    return buf.getvalue()


# ── 도매 본인 상품 관리 내보내기(내부용 — QR/가격셰이핑 없는 원장) ──────────────
PRODUCT_HEADERS = ["품번", "플랫폼코드", "상품명", "분류", "색상", "사이즈",
                   "도매가", "판매가", "재고", "혼용률"]


def products_xlsx_bytes(products: list[dict]) -> bytes:
    """도매 본인 상품 목록 → SKU 단위 1행 엑셀. 관리뷰(도매가+판매가) 그대로."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "상품목록"
    ws.append(PRODUCT_HEADERS)
    for p in products:
        skus = p.get("skus") or [{}]
        for s in skus:
            ws.append([
                p.get("source_p_number"), p.get("platform_code"), p.get("item_name"),
                p.get("category"), s.get("color"), s.get("size"),
                s.get("wholesale_price"), s.get("retail_price"), s.get("stock"),
                p.get("fabric_composition"),
            ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── 스타일 렌더 엑셀(사진·QR 박은 A~L) — jinsup excel_builder 1:1 흡수 ───────────────
# 컬럼: A 사진·B 품번·C 상품명·D 색상·E 상세사이즈·F 혼용률·G 도매가·H 판매가·I 재고·J P CODE·K QR 링크·L QR 이미지
# K(QR 링크) = QR 타깃 URL 텍스트, L(QR 이미지) = 실제 QR PNG 임베드. 둘 다 항상 채운다(역할 무관).
# 행 단위: **SKU(색상×사이즈)당 1행, 셀 병합 없이 매 행 독립** — 같은 사진/품번/상품명/QR 이라도
# 모든 행에 반복(사용자 결정 2026-06-06). jinsup _write_rows 방식.
RENDER_SHEET = "방송제품목록"
RENDER_HEADERS = ["사진", "품번", "상품명", "색상", "상세사이즈", "혼용률",
                  "도매가", "판매가", "재고", "P CODE", "QR 링크", "QR 이미지"]
_LEFT_ALIGN = {3, 5, 6, 11}         # 좌정렬 열(1-base): 상품명·상세사이즈·혼용률·QR 링크(URL 텍스트)
_MONEY_COLS = {7, 8}               # 콤마 서식(#,##0): 도매가·판매가
_IMG_COL, _QR_COL = "A", "L"       # 사진=A, QR 이미지=L
EXCEL_CELL_PX = 110                 # A열 사진 박스(px) — 행높이 95pt(~127px)·열폭 16(~112px) 안에 맞음
QR_PX = 100                         # QR 크기 (jinsup 보존)
_ROW_H, _HEADER_H = 95, 28


def _price_cell(v):
    """가격 셀 값 — None(미노출/미설정)은 빈 칸. 숫자는 그대로(콤마는 셀 number_format 이 처리)."""
    return v if isinstance(v, (int, float)) else ""


def price_code(wholesale, retail) -> str:
    """가격코드 = '도매가 앞2자리_판매가 앞2자리' (예: 18000/29000 → '18_29', 고가 3200000 → '32').

    도매가·판매가가 둘 다 숫자일 때만 생성(한쪽만/None 이면 빈칸).
    (jinsup 레거시는 //1000 였으나 고가 상품에서 코드가 길어져 '앞 2자리' 규칙으로 — 사용자 확정 2026-06-07.
     ⚠️ 호출부 책임: 가격 완전 미노출(none) 셀러는 raw 가격을 넘기지 말고 빈 p_code 를 줄 것.)
    """
    if isinstance(wholesale, (int, float)) and isinstance(retail, (int, float)):
        return f"{str(int(wholesale))[:2]}_{str(int(retail))[:2]}"
    return ""


_p_code = price_code  # 하위호환 별칭(_write_render_rows 폴백 등)


def cell_image_path(img: dict | None) -> str | None:
    """엑셀 셀용 이미지 경로 — **썸네일 우선(가벼움), 없으면 원본 fallback**.

    대량 업로드는 서버가 thumbnail_path(긴 변 800px)를 자동 생성한다. 셀은 어차피 110px 로
    렌더되므로 원본 대신 썸네일을 받으면 화질 차이 없이 Storage egress 를 줄인다(audit Critical).
    단일 업로드 등 썸네일이 없으면 원본(storage_path)으로 fallback.
    """
    if not img:
        return None
    return img.get("thumbnail_path") or img.get("storage_path")


# 인스턴스 단위 TTL 캐시 — 같은 경로의 셀 이미지 재다운로드 방지(반복 export 비용↓, audit Critical).
# 성공 바이트만 캐시(실패는 재시도 허용). Cloud Run 인스턴스/프로세스 내에서만 유효(cold start 시 비움).
_CELL_CACHE: "OrderedDict[str, tuple[float, bytes]]" = OrderedDict()
_CELL_CACHE_MAX = 512
_CELL_CACHE_TTL = 600  # 10분 — 같은 경로 이미지가 교체되면 그 사이 stale 가능(B2B 라 드묾)


def _cache_get(path: str) -> bytes | None:
    hit = _CELL_CACHE.get(path)
    if not hit:
        return None
    if time.time() - hit[0] >= _CELL_CACHE_TTL:
        _CELL_CACHE.pop(path, None)
        return None
    _CELL_CACHE.move_to_end(path)
    return hit[1]


def _cache_put(path: str, data: bytes) -> None:
    _CELL_CACHE[path] = (time.time(), data)
    _CELL_CACHE.move_to_end(path)
    while len(_CELL_CACHE) > _CELL_CACHE_MAX:
        _CELL_CACHE.popitem(last=False)


def cell_image_bytes(sb, storage_path: str | None, bucket: str = "product-images") -> bytes | None:
    """Storage 이미지 다운로드 → 셀용 썸네일 JPEG. 없음/실패 시 None(빌더가 '사진 없음' 처리).

    한 장씩만 메모리에 올렸다가 즉시 축소 → 원본 바이트는 곧 회수(대량 출력 OOM 방지).
    같은 경로는 TTL 캐시로 재다운로드를 건너뛴다(반복 export egress·지연↓).
    """
    if not storage_path:
        return None
    cached = _cache_get(storage_path)
    if cached is not None:
        return cached
    try:
        raw = sb.storage.from_(bucket).download(storage_path)
    except Exception:
        log.warning("엑셀 셀 이미지 다운로드 실패 path=%s", storage_path)
        return None
    res = process_image_bytes(raw, box=(EXCEL_CELL_PX, EXCEL_CELL_PX))
    if res.status != "ok" or not res.data:
        return None
    _cache_put(storage_path, res.data)   # 성공만 캐시
    return res.data


def _embed_image(ws: Worksheet, raw: bytes, anchor: str, max_px: int) -> None:
    img = XLImage(io.BytesIO(raw))
    if img.width and img.height and (img.width > max_px or img.height > max_px):
        ratio = min(max_px / img.width, max_px / img.height)
        img.width = int(img.width * ratio)
        img.height = int(img.height * ratio)
    ws.add_image(img, anchor)


def _write_render_rows(ws: Worksheet, rows: list[dict], base_url: str) -> None:
    """SKU당 1행 — 병합 없이 매 행 독립. 상품 공통(사진·품번·상품명·혼용률·QR)은 같아도 매 행 반복.

    K열(QR 링크) = QR 타깃 URL(텍스트), L열(QR 이미지) = 실제 QR PNG 임베드 — 둘 다 항상 채운다(역할 무관).
    P CODE(J) 는 호출부가 raw 가격으로 계산해 sku["p_code"] 로 주면 그걸 쓰고, 없으면 셰이핑된
    도매·판매로 폴백(_p_code, 기존 동작 보존).
    """
    r = 2
    for p in rows:
        skus = p.get("skus") or [{}]
        img_bytes = p.get("image_bytes")
        code = p.get("platform_code")
        qr_url = qr_target_url(code, base_url) if code else None
        qr_png = None
        if code:                                  # QR PNG 생성(상품당 1회, 행마다 재사용)
            try:
                qr_png = generate_qr_png(qr_url)
            except Exception:
                qr_png = None
        for s in skus:
            w, ret = s.get("wholesale_price"), s.get("retail_price")
            # 상품 공통 — 매 행 반복
            ws.cell(r, 2, p.get("source_p_number"))
            ws.cell(r, 3, p.get("item_name"))
            ws.cell(r, 6, p.get("fabric_composition"))
            # SKU 단위
            ws.cell(r, 4, s.get("color"))
            ws.cell(r, 5, s.get("size"))
            ws.cell(r, 7, _price_cell(w))
            ws.cell(r, 8, _price_cell(ret))
            ws.cell(r, 9, s.get("stock"))
            ws.cell(r, 10, s["p_code"] if "p_code" in s else _p_code(w, ret))  # J P CODE
            # 사진(A) — 같은 사진이라도 매 행 삽입
            if img_bytes:
                _embed_image(ws, img_bytes, f"A{r}", EXCEL_CELL_PX)
            else:
                ws.cell(r, 1, "사진 없음")
            # K(QR 링크) = URL 텍스트 — 항상 기록
            if qr_url:
                ws.cell(r, 11, qr_url)
            # L(QR 이미지) = QR PNG 임베드 — 항상 기록
            if qr_png is not None:
                _embed_image(ws, qr_png, f"L{r}", QR_PX)
            elif code:
                ws.cell(r, 12, "QR 오류")
            r += 1


def _apply_render_styles(ws: Worksheet) -> None:
    font_main = Font(name="맑은 고딕", size=10)
    font_head = Font(name="맑은 고딕", size=11, bold=True)
    fill_head = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    side = Side(style="thin", color="D3D3D3")
    border = Border(left=side, right=side, top=side, bottom=side)

    ws.row_dimensions[1].height = _HEADER_H
    for c in ws[1]:
        c.font = font_head
        c.fill = fill_head
        c.alignment = center
        c.border = border

    ncols = len(RENDER_HEADERS)
    max_row = ws.max_row
    for ri in range(2, max_row + 1):
        ws.row_dimensions[ri].height = _ROW_H
        for ci in range(1, ncols + 1):
            cell = ws.cell(ri, ci)
            cell.font = font_main
            cell.border = border
            cell.alignment = left if ci in _LEFT_ALIGN else center  # K(QR 링크)는 _LEFT_ALIGN 포함
            if ci in _MONEY_COLS:
                cell.number_format = "#,##0"   # 도매가·판매가 콤마 표기(예: 18,000)

    ws.column_dimensions["A"].width = 16            # 사진
    ws.column_dimensions["K"].width = 48            # QR 링크(URL 텍스트) — 넓게
    ws.column_dimensions["L"].width = 16            # QR 이미지 — 이미지 박스폭
    for ci in range(2, 11):  # B~J 자동폭(내용 길이 기준)
        letter = ws.cell(1, ci).column_letter
        max_len = max((len(str(ws.cell(ri, ci).value or "")) for ri in range(1, max_row + 1)), default=0)
        ws.column_dimensions[letter].width = max(max_len + 4, 13)


def build_render_xlsx(rows: list[dict], base_url: str) -> bytes:
    """사진·QR 박은 스타일 카탈로그 엑셀(A~L) 바이트.

    가격은 **호출부에서 visible_price() 로 셰이핑 완료된 값**을 받는다(빌더는 셰이핑/조회 안 함).
    rows[i] = {
        "source_p_number": str, "item_name": str, "fabric_composition": str | None,
        "platform_code": str, "image_bytes": bytes | None,
        "skus": [{"color", "size", "stock", "wholesale_price"|None, "retail_price"|None,
                  "p_code"?: str}, ...],   # p_code 주면 J열에 그대로, 없으면 도매·판매로 폴백
    }  (빈 skus 면 1행으로, 색상/사이즈/가격 공백)

    QR 은 두 열 — K열(QR 링크) = QR 타깃 URL 텍스트, L열(QR 이미지) = 실제 QR PNG 임베드. 둘 다 항상 출력.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = RENDER_SHEET
    ws.append(RENDER_HEADERS)
    _write_render_rows(ws, rows, base_url)
    _apply_render_styles(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
