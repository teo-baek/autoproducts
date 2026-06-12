import csv
import openpyxl
import pytest
from app.services.excel_parse import ExcelFormatError, parse_template_rows, TEMPLATE_COLUMNS

def _make_xlsx(tmp_path, rows):
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(TEMPLATE_COLUMNS)
    for r in rows: ws.append(r)
    p = tmp_path / "t.xlsx"; wb.save(p); return p

def _make_csv(tmp_path, rows, encoding="utf-8-sig"):
    p = tmp_path / "t.csv"
    with open(p, "w", newline="", encoding=encoding) as f:
        w = csv.writer(f)
        w.writerow(TEMPLATE_COLUMNS)
        for r in rows: w.writerow(r)
    return p

def _make_xls(tmp_path, rows):
    import xlwt
    wb = xlwt.Workbook(); ws = wb.add_sheet("s")
    for j, c in enumerate(TEMPLATE_COLUMNS): ws.write(0, j, c)
    for i, r in enumerate(rows, start=1):
        for j, v in enumerate(r): ws.write(i, j, v)
    p = tmp_path / "t.xls"; wb.save(str(p)); return p


def test_parse_csv_valid(tmp_path):
    p = _make_csv(tmp_path, [["1001", "린넨셔츠", "화이트", "F", "12000", "29000"]])
    parsed = parse_template_rows(str(p))
    assert parsed.errors == []
    assert parsed.rows[0]["source_p_number"] == "1001"
    assert parsed.rows[0]["wholesale_price"] == 12000

def test_parse_csv_cp949_korean(tmp_path):
    # 한글 엑셀의 ANSI(cp949) CSV 도 읽혀야 함
    p = _make_csv(tmp_path, [["1002", "한글상품", "블랙", "L", "15000", "30000"]], encoding="cp949")
    parsed = parse_template_rows(str(p))
    assert parsed.rows[0]["item_name"] == "한글상품"

def test_parse_currency_and_won(tmp_path):
    # 도매가/판매가에 ₩·원·콤마가 섞여도 숫자로 해석 ('분명 맞는데' 케이스)
    p = _make_xlsx(tmp_path, [["F-1", "실크블라우스", "블랙", "M", "₩280,000원", "350,000 원"]])
    parsed = parse_template_rows(str(p))
    assert parsed.errors == []
    assert parsed.rows[0]["wholesale_price"] == 280000
    assert parsed.rows[0]["retail_price"] == 350000


def test_parse_header_order_independent(tmp_path):
    # 열 순서가 다르고 추가 열(카테고리)이 있어도 헤더 이름으로 매칭
    p = tmp_path / "t.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["상품명", "품번", "카테고리", "도매가", "판매가", "색상", "사이즈"])
    ws.append(["실크블라우스", "F-SLK-001", "의류", "280000", "350000", "아이보리", "S"])
    wb.save(p)
    parsed = parse_template_rows(str(p))
    assert parsed.errors == []
    r = parsed.rows[0]
    assert r["source_p_number"] == "F-SLK-001"
    assert r["wholesale_price"] == 280000
    assert r["color"] == "아이보리"


def test_parse_alias_header(tmp_path):
    # 별칭 헤더(공급가/소비자가/상품코드)도 인식
    p = tmp_path / "t.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["상품코드", "상품명", "색상", "사이즈", "공급가", "소비자가"])
    ws.append(["A-1", "코트", "네이비", "L", "120000", "200000"])
    wb.save(p)
    parsed = parse_template_rows(str(p))
    assert parsed.errors == []
    assert parsed.rows[0]["wholesale_price"] == 120000


def test_parse_color_alias_kalra(tmp_path):
    # 실데이터(20260608opn.xls): 색상 헤더가 '칼라', 상품명이 '품명', 품번은 전부 빈칸.
    # '칼라'를 색상으로 인식 + 품번 없으면 이지머스 코드 등록 → 정상 파싱돼야 한다.
    p = tmp_path / "t.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["품번", "품명", "칼라", "사이즈", "입고가", "도매가", "소매가"])
    ws.append(["", "레이온타이나시BL", "소라", "FREE", 0, 14000, 0])
    wb.save(p)
    parsed = parse_template_rows(str(p))
    assert parsed.errors == []
    assert len(parsed.rows) == 1
    r = parsed.rows[0]
    assert r["color"] == "소라"
    assert r["item_name"] == "레이온타이나시BL"
    assert r["wholesale_price"] == 14000          # 입고가(0) 아닌 도매가(14000) 선택
    assert r["source_p_number"] is None           # 품번 없음 → 이지머스 코드 대상


def test_parse_blank_pnum_with_name_kept_for_auto_code(tmp_path):
    # 품번 없지만 상품명 있으면 → 폐기 X. 진짜 상품으로 보고 '이지머스 자체 품번'으로 등록(QA 3차 1p).
    # 파서는 source_p_number=None 마커만 남기고(ingest 가 platform_code 로 채움), 오류·드롭 아님.
    p = _make_xlsx(tmp_path, [["", "프릴리OPS", "블랙", "F", "18000", ""]])
    parsed = parse_template_rows(str(p))
    assert parsed.dropped == 0
    assert parsed.errors == []
    assert len(parsed.rows) == 1
    assert parsed.rows[0]["source_p_number"] is None     # 이지머스 코드 부여 대상 마커
    assert parsed.rows[0]["item_name"] == "프릴리OPS"


def test_parse_blank_pnum_no_name_dropped(tmp_path):
    # 품번·상품명 둘 다 없는 행(POS 합계/소계 잡행) → 조용히 폐기(dropped, 오류 아님)
    p = _make_xlsx(tmp_path, [["", "", "", "", "999000", ""]])
    parsed = parse_template_rows(str(p))
    assert parsed.rows == []
    assert parsed.dropped == 1
    assert parsed.errors == []


def test_parse_blank_pnum_with_name_registers_alongside_valid(tmp_path):
    # 품번 있는 행 + 품번 없지만 상품명 있는 행 → 둘 다 등록 대상(후자는 이지머스 코드)
    p = _make_xlsx(tmp_path, [
        ["1001", "셔츠", "화이트", "F", "12000", "29000"],
        ["", "품번없음상품", "블랙", "F", "18000", ""],
    ])
    parsed = parse_template_rows(str(p))
    assert parsed.dropped == 0 and parsed.errors == []
    assert len(parsed.rows) == 2
    assert parsed.rows[0]["source_p_number"] == "1001"
    assert parsed.rows[1]["source_p_number"] is None     # 이지머스 코드 부여 대상


def test_parse_xls_valid(tmp_path):
    # 구형 엑셀(.xls) — 숫자는 float 로 들어와도 _to_int 가 보정
    p = _make_xls(tmp_path, [["2001", "울코트", "네이비", "M", 50000, 99000]])
    parsed = parse_template_rows(str(p))
    assert parsed.errors == []
    assert parsed.rows[0]["wholesale_price"] == 50000
    assert parsed.rows[0]["retail_price"] == 99000

def test_parse_valid_rows(tmp_path):
    p = _make_xlsx(tmp_path, [["1001","린넨셔츠","화이트","F","12000","29000"]])
    parsed = parse_template_rows(str(p))
    assert parsed.rows[0]["source_p_number"] == "1001"
    assert parsed.rows[0]["wholesale_price"] == 12000
    assert parsed.errors == []

def test_parse_missing_price_is_error(tmp_path):
    p = _make_xlsx(tmp_path, [["1002","바지","블랙","L","","20000"]])
    parsed = parse_template_rows(str(p))
    assert parsed.errors and parsed.errors[0]["row"] == 2
    assert parsed.errors[0]["field"] == "도매가"          # 필드 단위 보고
    assert parsed.errors[0]["reason"] == "필수 값이 누락되었습니다"


def test_parse_non_numeric_price_is_error(tmp_path):
    p = _make_xlsx(tmp_path, [["1003","니트","그레이","M","열두개","20000"]])
    parsed = parse_template_rows(str(p))
    assert parsed.rows == []
    e = parsed.errors[0]
    assert e["field"] == "도매가" and e["reason"] == "숫자 형식이 아닙니다"


def test_parse_missing_item_name_reports_field(tmp_path):
    # 품번은 있는데 상품명만 없으면 → 상품명 누락 오류 (품번 없는 건 폐기 처리라 별도)
    p = _make_xlsx(tmp_path, [["1001", "", "블랙", "L", "10000", ""]])
    parsed = parse_template_rows(str(p))
    fields = {e["field"] for e in parsed.errors}
    assert "상품명" in fields


def test_parse_blank_color_is_error(tmp_path):
    # 색상 빈칸 → 검증 오류(필드 '색상'). DB product_skus.color NOT NULL 과 일치시켜 커밋 실패를 사전 차단.
    p = _make_xlsx(tmp_path, [["1001", "셔츠", "", "F", "12000", "29000"]])
    parsed = parse_template_rows(str(p))
    assert parsed.rows == []                       # 누락 행은 등록 대상에서 제외
    assert "색상" in {e["field"] for e in parsed.errors}
    assert parsed.errors[0]["reason"] == "필수 값이 누락되었습니다"


def test_parse_blank_size_is_error(tmp_path):
    # 사이즈 빈칸 → 검증 오류(필드 '사이즈')
    p = _make_xlsx(tmp_path, [["1001", "셔츠", "화이트", "", "12000", "29000"]])
    parsed = parse_template_rows(str(p))
    assert parsed.rows == []
    assert "사이즈" in {e["field"] for e in parsed.errors}


def test_parse_numeric_size_normalized_to_text(tmp_path):
    # 숫자 사이즈(95)는 '95' 텍스트로 정규화 — '95.0' 부동소수 잔재 방지(TEXT 컬럼 안전)
    p = _make_xls(tmp_path, [["2001", "팬츠", "네이비", 95, 50000, 99000]])
    parsed = parse_template_rows(str(p))
    assert parsed.errors == []
    assert parsed.rows[0]["size"] == "95"


def test_parse_missing_color_size_column_reports_row_errors(tmp_path):
    # 색상·사이즈 컬럼이 아예 없어도 '파일 자체를 막지 않고', 행별로 '색상/사이즈 누락'을
    # 검증 표(errors)에 보여준다 — 어디가 빠졌는지 정확히(사용자 피드백).
    p = tmp_path / "t.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["품번", "상품명", "도매가", "판매가"])      # 색상/사이즈 없음
    ws.append(["1001", "셔츠", "12000", "29000"])
    wb.save(p)
    parsed = parse_template_rows(str(p))                 # 예외 없이 통과
    fields = {e["field"] for e in parsed.errors}
    assert "색상" in fields and "사이즈" in fields
    assert parsed.rows == []                             # 등록 대상엔 안 들어감
    assert all(e["row"] == 2 for e in parsed.errors)     # 해당 행 번호로 표에 표시


def test_parse_blank_trailing_row_skipped(tmp_path):
    p = _make_xlsx(tmp_path, [["1001","셔츠","화이트","F","12000","29000"], [None]*6])
    parsed = parse_template_rows(str(p))
    assert len(parsed.rows) == 1 and parsed.errors == []   # 빈 행은 오류 아님


# ── 동의어 헤더 확장(jinsup SYNONYMS_POOL 흡수) ──────────────────────────────
def test_parse_extended_aliases(tmp_path):
    # 모델명→품번, 물품명→상품명, 색상명→색상, 상세사이즈→사이즈, 도매단가→도매가, 매장판매가→판매가
    p = tmp_path / "t.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["모델명", "물품명", "색상명", "상세사이즈", "도매단가", "매장판매가"])
    ws.append(["M-9", "트렌치", "카멜", "55", "88000", "159000"])
    wb.save(p)
    parsed = parse_template_rows(str(p))
    assert parsed.errors == []
    r = parsed.rows[0]
    assert r["source_p_number"] == "M-9"
    assert r["item_name"] == "트렌치"
    assert r["color"] == "카멜"
    assert r["size"] == "55"
    assert r["wholesale_price"] == 88000
    assert r["retail_price"] == 159000


def test_wholesale_priority_when_multiple_price_columns(tmp_path):
    # 실제 POS 처럼 입고가·도매가·도매Sale 이 동시에 있으면 '도매가'(우선순위 앞) 를 택함
    p = tmp_path / "t.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["품번", "상품명", "색상", "사이즈", "입고가", "도매가", "도매Sale", "소매가"])
    ws.append(["1001", "코트", "블랙", "F", "10000", "18000", "15000", "39000"])
    wb.save(p)
    parsed = parse_template_rows(str(p))
    assert parsed.errors == []
    assert parsed.rows[0]["wholesale_price"] == 18000     # 입고가(10000)/도매Sale(15000) 아님


def test_stock_column_ingested(tmp_path):
    # 재고정상 → stock 인입(관대 파싱). 혼용률 → fabric_composition.
    p = tmp_path / "t.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["품번", "상품명", "색상", "사이즈", "도매가", "판매가", "혼용률", "재고정상"])
    ws.append(["1001", "니트", "그레이", "M", "12000", "29000", "울 80% 나일론 20%", "7"])
    wb.save(p)
    parsed = parse_template_rows(str(p))
    assert parsed.errors == []
    r = parsed.rows[0]
    assert r["stock"] == 7
    assert r["fabric_composition"] == "울 80% 나일론 20%"


def test_stock_non_numeric_is_lenient(tmp_path):
    # 재고는 선택 메타 — 비숫자여도 행을 막지 않고 0 으로(가격과 다른 정책)
    p = tmp_path / "t.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["품번", "상품명", "색상", "사이즈", "도매가", "재고"])
    ws.append(["1001", "셔츠", "화이트", "F", "12000", "품절"])
    wb.save(p)
    parsed = parse_template_rows(str(p))
    assert parsed.errors == []
    assert parsed.rows[0]["stock"] == 0


def test_missing_required_column_raises_file_error(tmp_path):
    # 도매가 컬럼이 헤더에 아예 없음 → 파일 단위 오류(행별로 흐려지지 않음)
    p = tmp_path / "t.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["품번", "상품명", "색상", "사이즈"])          # 도매가 없음
    ws.append(["1001", "셔츠", "화이트", "F"])
    wb.save(p)
    with pytest.raises(ExcelFormatError) as ei:
        parse_template_rows(str(p))
    assert "도매가" in str(ei.value)


def test_missing_all_required_columns_raises_file_error(tmp_path):
    # 표준과 전혀 다른 헤더(인식 불가) → 파일 단위 오류(위치 폴백 안 함)
    p = tmp_path / "t.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["메모", "비고", "기타"])
    ws.append(["a", "b", "c"])
    wb.save(p)
    with pytest.raises(ExcelFormatError):
        parse_template_rows(str(p))
