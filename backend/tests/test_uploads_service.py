import openpyxl
import pytest
from app.services.uploads import (
    ingest_excel, attach_images, preview_excel, stage_zip_to_manifest, resolve_match,
    list_unmatched, UploadError, UploadForbidden,
)


class FakeUploadRepo:
    def __init__(self):
        self.products = []; self.skus = []; self.jobs = []; self.images = []
        self.seq = 0; self._pmap = {}

    def next_platform_code(self):
        self.seq += 1; return f"EZM-{self.seq:06d}"

    def insert_product(self, d):
        d = {**d, "id": f"p{len(self.products) + 1}"}
        self.products.append(d); self._pmap[d["source_p_number"]] = d["id"]; return d

    def insert_skus(self, rows):
        self.skus.extend(rows); return rows

    def create_upload_job(self, d):
        d = {**d, "id": "job-1"}; self.jobs.append(d); return d

    def update_upload_job(self, jid, patch):
        for j in self.jobs:
            if j["id"] == jid: j.update(patch); return j
        return {"id": jid, **patch}

    def get_upload_job(self, jid):
        return next((j for j in self.jobs if j["id"] == jid), None)   # 없는 job → None

    def products_pnum_map(self, wid):
        return dict(self._pmap)

    def insert_images(self, rows):
        out = []
        for r in rows:
            r = {**r, "id": f"img{len(self.images) + 1}"}; self.images.append(r); out.append(r)
        return out

    def list_unmatched_images(self, wid):
        return [i for i in self.images if i["match_status"] == "unmatched"]

    def update_image(self, iid, patch, wholesaler_id=None):
        for i in self.images:
            if i["id"] == iid and (wholesaler_id is None or i.get("wholesaler_id") == wholesaler_id):
                i.update(patch); return i
        return None


def _make_xlsx(path, rows):
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["품번", "상품명", "색상", "사이즈", "도매가", "판매가"])
    for r in rows:
        ws.append(r)
    wb.save(path)


def test_ingest_excel_groups_rows_into_products_and_skus(tmp_path):
    p = tmp_path / "in.xlsx"
    _make_xlsx(p, [
        ("1001", "린넨셔츠", "화이트", "F", 12000, 29000),
        ("1001", "린넨셔츠", "블랙", "F", 12000, 29000),   # 같은 품번 → 같은 상품, sku 추가
        ("1002", "데님", "인디고", "M", 20000, 45000),
    ])
    repo = FakeUploadRepo()
    out = ingest_excel(repo, "w1", str(p), created_by="staff-1")
    assert len(out["products"]) == 2            # 품번 2종 → 상품 2개
    assert len(repo.skus) == 3                  # sku 3행
    assert out["job"]["status"] == "needs_matching"
    assert out["job"]["total_rows"] == 3
    assert repo.products[0]["created_by"] == "staff-1"


def test_ingest_excel_noncontiguous_same_pnum_splits(tmp_path):
    # 품번은 유일키가 아님 — 같은 품번이 떨어져 다시 나오면 별개 상품으로 분리(연속 블록 그룹핑)
    p = tmp_path / "in.xlsx"
    _make_xlsx(p, [
        ("915", "블라우스A", "화이트", "F", 10000, 20000),
        ("700", "팬츠", "블랙", "M", 15000, 30000),
        ("915", "원피스B", "레드", "F", 18000, 36000),     # 같은 품번 915, 떨어져 등장 → 별개
    ])
    out = ingest_excel(FakeUploadRepo(), "w1", str(p))
    assert len(out["products"]) == 3
    assert [pr["item_name"] for pr in out["products"]] == ["블라우스A", "팬츠", "원피스B"]


def test_ingest_excel_adjacent_same_pnum_diff_name_splits(tmp_path):
    # 인접해도 상품명이 다르면 별개 상품
    p = tmp_path / "in.xlsx"
    _make_xlsx(p, [
        ("915", "블라우스A", "화이트", "F", 10000, 20000),
        ("915", "원피스B", "레드", "F", 18000, 36000),
    ])
    out = ingest_excel(FakeUploadRepo(), "w1", str(p))
    assert len(out["products"]) == 2


def test_ingest_excel_contiguous_same_pnum_name_groups(tmp_path):
    # 연속된 같은 (품번, 상품명) = 한 상품의 여러 SKU
    p = tmp_path / "in.xlsx"
    _make_xlsx(p, [
        ("915", "블라우스", "화이트", "F", 10000, 20000),
        ("915", "블라우스", "블랙", "F", 10000, 20000),
    ])
    repo = FakeUploadRepo()
    out = ingest_excel(repo, "w1", str(p))
    assert len(out["products"]) == 1 and len(repo.skus) == 2


def test_ingest_excel_records_parse_errors(tmp_path):
    p = tmp_path / "in.xlsx"
    _make_xlsx(p, [
        ("1001", "정상", "화이트", "F", 12000, 29000),
        ("1002", "", "블랙", "F", 12000, 29000),            # 상품명 누락 → error
        ("1003", "가격이상", "레드", "F", "NOTNUM", 29000),  # 도매가 정수변환 실패 → error
    ])
    repo = FakeUploadRepo()
    out = ingest_excel(repo, "w1", str(p))
    assert len(out["products"]) == 1
    assert out["job"]["error_rows"] == 2
    assert len(out["errors"]) == 2


def test_ingest_excel_no_valid_rows_marks_failed(tmp_path):
    p = tmp_path / "in.xlsx"
    # 품번·상품명 둘 다 없으면 식별 불가 → 유효행 0 → failed
    _make_xlsx(p, [("", "", "화이트", "F", 12000, 29000)])
    out = ingest_excel(FakeUploadRepo(), "w1", str(p))
    assert out["products"] == []
    assert out["job"]["status"] == "failed"


def test_attach_images_matches_by_filename_and_updates_job():
    repo = FakeUploadRepo()
    repo.create_upload_job({"wholesaler_id": "w1"})
    repo.insert_product({"wholesaler_id": "w1", "source_p_number": "1001",
                         "platform_code": "EZM-1", "item_name": "셔츠"})
    out = attach_images(repo, "job-1", [
        {"original_filename": "1001_front.jpg", "storage_path": "w1/1001_front.jpg"},
        {"original_filename": "9999.jpg", "storage_path": "w1/9999.jpg"},
    ], created_by="staff-1", caller_wid="w1")
    assert out["matched"] == ["1001_front.jpg"]
    assert out["unmatched"] == ["9999.jpg"]
    assert any(i["match_status"] == "matched" and i["product_id"] == "p1" for i in repo.images)
    job = repo.get_upload_job("job-1")
    assert job["matched_rows"] == 1
    assert job["status"] == "needs_matching"     # 미매칭 잔존 → 계속 매칭 대기


def test_attach_images_all_matched_completes_job():
    repo = FakeUploadRepo()
    repo.create_upload_job({"wholesaler_id": "w1"})
    repo.insert_product({"wholesaler_id": "w1", "source_p_number": "1001",
                         "platform_code": "EZM-1", "item_name": "셔츠"})
    attach_images(repo, "job-1", [{"original_filename": "1001.jpg", "storage_path": "w1/1001.jpg"}],
                  caller_wid="w1")
    assert repo.get_upload_job("job-1")["status"] == "completed"


def test_resolve_match_links_image_to_product():
    repo = FakeUploadRepo()
    repo.create_upload_job({"wholesaler_id": "w1"})
    repo.insert_product({"wholesaler_id": "w1", "source_p_number": "1001",
                         "platform_code": "EZM-1", "item_name": "셔츠"})
    repo.insert_images([{"wholesaler_id": "w1", "storage_path": "w1/x.jpg",
                         "original_filename": "x.jpg", "product_id": None, "match_status": "unmatched"}])
    out = resolve_match(repo, "job-1", image_id="img1", source_p_number="1001", caller_wid="w1")
    assert out["product_id"] == "p1" and out["match_status"] == "matched"


def test_resolve_match_unknown_pnum_raises():
    repo = FakeUploadRepo()
    repo.create_upload_job({"wholesaler_id": "w1"})
    with pytest.raises(UploadError):
        resolve_match(repo, "job-1", image_id="img1", source_p_number="NOPE", caller_wid="w1")


def test_uploads_reject_foreign_caller_idor():
    """도매 w1 의 job 을 w2 가 건드리면 모두 404(UploadForbidden) — IDOR 차단."""
    repo = FakeUploadRepo()
    repo.create_upload_job({"wholesaler_id": "w1"})
    repo.insert_product({"wholesaler_id": "w1", "source_p_number": "1001",
                         "platform_code": "EZM-1", "item_name": "셔츠"})
    with pytest.raises(UploadForbidden):
        attach_images(repo, "job-1", [{"original_filename": "1001.jpg", "storage_path": "x"}], caller_wid="w2")
    with pytest.raises(UploadForbidden):
        list_unmatched(repo, "job-1", caller_wid="w2")
    with pytest.raises(UploadForbidden):
        resolve_match(repo, "job-1", image_id="img1", source_p_number="1001", caller_wid="w2")


def test_resolve_match_foreign_image_forbidden():
    """job 은 내 소유여도, 대상 이미지가 타 업체 소유면 갱신 거부."""
    repo = FakeUploadRepo()
    repo.create_upload_job({"wholesaler_id": "w1"})
    repo.insert_product({"wholesaler_id": "w1", "source_p_number": "1001",
                         "platform_code": "EZM-1", "item_name": "셔츠"})
    repo.insert_images([{"wholesaler_id": "w2", "storage_path": "w2/x.jpg",   # 타 업체 이미지
                         "original_filename": "x.jpg", "product_id": None, "match_status": "unmatched"}])
    with pytest.raises(UploadForbidden):
        resolve_match(repo, "job-1", image_id="img1", source_p_number="1001", caller_wid="w1")


def test_ingest_excel_duplicate_pnum_becomes_error_not_crash():
    """재업로드(품번 UNIQUE 충돌)는 해당 품번만 error 로 떨구고 나머지는 계속."""
    class DupRepo(FakeUploadRepo):
        def insert_product(self, d):
            if d["source_p_number"] == "DUP":
                raise Exception("duplicate key value violates unique constraint")
            return super().insert_product(d)
    import openpyxl as _x
    import tempfile, os
    fd, p = tempfile.mkstemp(suffix=".xlsx"); os.close(fd)
    wb = _x.Workbook(); ws = wb.active
    ws.append(["품번", "상품명", "색상", "사이즈", "도매가", "판매가"])
    ws.append(("DUP", "중복", "화이트", "F", 1000, 2000))
    ws.append(("OK1", "정상", "블랙", "F", 1000, 2000))
    wb.save(p)
    out = ingest_excel(DupRepo(), "w1", p)
    os.unlink(p)
    assert len(out["products"]) == 1                      # OK1 만 생성
    assert any(e.get("source_p_number") == "DUP" for e in out["errors"])
    assert out["job"]["status"] == "needs_matching"       # 일부라도 생성됨


class FakeStorageRepo(FakeUploadRepo):
    """원본을 다운로드/가공/업로드하는 가공 경로 검증용 — 인메모리 Storage."""
    def __init__(self):
        super().__init__()
        self.store = {}                       # path -> bytes

    def put(self, path, data):
        self.store[path] = data

    def download_object(self, path, bucket="product-images"):
        return self.store[path]               # 없으면 KeyError → 서비스가 'none' 처리

    def upload_object(self, path, data, bucket="product-images", content_type="image/jpeg"):
        self.store[path] = data
        return path


def _jpeg_bytes(w=1200, h=900):
    import io
    from PIL import Image
    img = Image.new("RGB", (w, h), (10, 120, 200))
    buf = io.BytesIO(); img.save(buf, format="JPEG"); return buf.getvalue()


def test_attach_images_generates_thumbnail_when_storage_available():
    repo = FakeStorageRepo()
    repo.create_upload_job({"wholesaler_id": "w1"})
    repo.insert_product({"wholesaler_id": "w1", "source_p_number": "1001",
                         "platform_code": "EZM-1", "item_name": "셔츠"})
    repo.put("w1/1001.jpg", _jpeg_bytes())     # 프론트가 올려둔 원본
    out = attach_images(repo, "job-1",
                        [{"original_filename": "1001.jpg", "storage_path": "w1/1001.jpg"}],
                        caller_wid="w1")
    assert out["processed"] == {"ok": 1, "none": 0, "error": 0}
    img = repo.images[0]
    assert img["thumbnail_path"] == "thumbs/w1/1001.jpg"     # 파생 썸네일 경로 기록
    assert "thumbs/w1/1001.jpg" in repo.store                # 실제 업로드됨
    assert repo.store["thumbs/w1/1001.jpg"][:2] == b"\xff\xd8"  # JPEG 산출물


def test_attach_images_missing_original_marked_none_not_crash():
    repo = FakeStorageRepo()
    repo.create_upload_job({"wholesaler_id": "w1"})
    repo.insert_product({"wholesaler_id": "w1", "source_p_number": "1001",
                         "platform_code": "EZM-1", "item_name": "셔츠"})
    # 원본을 store 에 넣지 않음 → download 실패 → 'none'(배치 안 죽음)
    out = attach_images(repo, "job-1",
                        [{"original_filename": "1001.jpg", "storage_path": "w1/1001.jpg"}],
                        caller_wid="w1")
    assert out["processed"]["none"] == 1
    assert repo.images[0]["thumbnail_path"] is None          # 가공 실패 → 원본 폴백


def test_attach_images_corrupt_original_marked_error():
    repo = FakeStorageRepo()
    repo.create_upload_job({"wholesaler_id": "w1"})
    repo.insert_product({"wholesaler_id": "w1", "source_p_number": "1001",
                         "platform_code": "EZM-1", "item_name": "셔츠"})
    repo.put("w1/1001.jpg", b"\x00 not an image \xff")        # 다운로드는 되나 가공 실패
    out = attach_images(repo, "job-1",
                        [{"original_filename": "1001.jpg", "storage_path": "w1/1001.jpg"}],
                        caller_wid="w1")
    assert out["processed"]["error"] == 1
    assert repo.images[0]["thumbnail_path"] is None


def test_attach_images_processes_unmatched_images_too():
    # 미매칭 이미지도 썸네일 생성(수동매칭 UI 프리뷰용)
    repo = FakeStorageRepo()
    repo.create_upload_job({"wholesaler_id": "w1"})
    repo.put("w1/9999.jpg", _jpeg_bytes())
    out = attach_images(repo, "job-1",
                        [{"original_filename": "9999.jpg", "storage_path": "w1/9999.jpg"}],
                        caller_wid="w1")
    assert out["unmatched"] == ["9999.jpg"]
    assert out["processed"]["ok"] == 1
    assert repo.images[0]["thumbnail_path"] == "thumbs/w1/9999.jpg"


def _zip_bytes(entries):
    import io
    import zipfile
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        for name, data in entries:
            zf.writestr(name, data)
    return buf.getvalue()


def test_stage_zip_uploads_and_returns_manifest():
    # 2단계 staging — 원본+썸네일을 Storage 에 올리고 매니페스트만 반환(상품/이미지/잡 미생성)
    repo = FakeStorageRepo()
    z = _zip_bytes([
        ("1001.jpg", _jpeg_bytes()),
        ("9999.png", _jpeg_bytes()),
        ("__MACOSX/._1001.jpg", b"junk"),         # 맥OS 잔재 → 제외
        ("readme.txt", b"hello"),                 # 비이미지 → 제외
    ])
    out = stage_zip_to_manifest(repo, "w1", z)
    assert out["processed"]["ok"] == 2
    names = {m["original_filename"] for m in out["manifest"]}
    assert names == {"1001.jpg", "9999.png"}
    m = next(m for m in out["manifest"] if m["original_filename"] == "1001.jpg")
    # 저장 키는 ASCII 안전 해시(원본명 그대로 아님) — staging 경로 + .jpg
    assert m["storage_path"].startswith("w1/staging/") and m["storage_path"].endswith(".jpg")
    assert m["storage_path"] in repo.store                          # 원본 staging 업로드
    assert m["thumbnail_path"] == "thumbs/" + m["storage_path"]     # 썸네일 생성
    assert m["thumbnail_path"] in repo.store
    assert repo.images == [] and repo.products == []     # DB 미기록(staging 단계)


def test_stage_zip_korean_filename_safe_key():
    # 한글 파일명도 InvalidKey 안 나게 ASCII 안전 키로 저장 + 원본명은 보존
    repo = FakeStorageRepo()
    out = stage_zip_to_manifest(repo, "w1", _zip_bytes([("수플레니트198.jpg", _jpeg_bytes())]))
    m = out["manifest"][0]
    assert m["original_filename"] == "수플레니트198.jpg"            # 원본 한글명 보존(매칭용)
    assert m["storage_path"].isascii()                              # 저장 키는 ASCII
    assert m["storage_path"].startswith("w1/staging/")


def test_stage_zip_no_images_raises():
    with pytest.raises(UploadError):
        stage_zip_to_manifest(FakeStorageRepo(), "w1", _zip_bytes([("readme.txt", b"x")]))


def test_stage_zip_bad_zip_raises():
    with pytest.raises(UploadError):
        stage_zip_to_manifest(FakeStorageRepo(), "w1", b"this is definitely not a zip")


def test_stage_zip_oversize_raises(monkeypatch):
    import app.services.uploads as up
    monkeypatch.setattr(up, "_ZIP_MAX_BYTES", 10)
    with pytest.raises(UploadError):
        up.stage_zip_to_manifest(FakeStorageRepo(), "w1", _zip_bytes([("1001.jpg", _jpeg_bytes())]))


def test_commit_flow_stage_then_attach_with_thumbnail():
    # 4단계 commit 흐름(서비스 단위): stage manifest → 상품 생성 → attach_images 가 staged 썸네일 재사용
    repo = FakeStorageRepo()
    z = _zip_bytes([("1001.jpg", _jpeg_bytes())])
    staged = stage_zip_to_manifest(repo, "w1", z)["manifest"]
    repo.create_upload_job({"wholesaler_id": "w1"})
    repo.insert_product({"wholesaler_id": "w1", "source_p_number": "1001",
                         "platform_code": "EZM-1", "item_name": "셔츠"})
    before = dict(repo.store)
    out = attach_images(repo, "job-1", staged, caller_wid="w1")
    assert out["matched"] == ["1001.jpg"]
    img = repo.images[0]
    assert img["product_id"] == "p1"
    assert img["thumbnail_path"] == staged[0]["thumbnail_path"]    # staged 썸네일 재사용
    assert out["processed"] == {"ok": 0, "none": 0, "error": 0}    # 재가공 안 함(이미 썸네일 있음)
    assert repo.store == before                                    # Storage 추가 변경 없음


def test_zip_member_name_recovers_korean():
    # Windows 제작 zip 의 한글 파일명(cp437 디코드 상태 + UTF-8 플래그 꺼짐) 복구
    import zipfile
    from app.services.uploads import _zip_member_name
    info = zipfile.ZipInfo()
    info.filename = "프릴원피스.jpg".encode("cp949").decode("cp437")  # Windows zip 저장 형태 모사
    info.flag_bits = 0                                               # UTF-8 플래그 꺼짐
    assert _zip_member_name(info) == "프릴원피스.jpg"


def test_preview_excel_dry_run(tmp_path):
    # 1단계 검증 — 카운트/오류/폐기수만, DB 미기록(repo 인자 자체가 없음)
    p = tmp_path / "prev.xlsx"
    _make_xlsx(p, [
        ("1001", "셔츠", "화이트", "F", 12000, 29000),
        ("1001", "셔츠", "블랙", "F", 12000, 29000),     # 같은 (품번,상품명) 연속 → 한 상품
        ("", "품번없음", "블랙", "F", 18000, ""),          # 품번 없음 → 폐기
    ])
    out = preview_excel(str(p))
    assert out["product_count"] == 1
    assert out["sku_count"] == 2
    assert out["dropped"] == 1
    assert out["errors"] == []


def test_insert_errors_friendly_duplicate_message():
    # 재업로드 중복(UNIQUE 충돌) → raw DB JSON 이 아니라 친화 사유로 기록
    class DupRepo(FakeUploadRepo):
        def insert_product(self, d):
            raise Exception(
                "{'message': 'duplicate key value violates unique constraint "
                "\"products_wholesaler_source_alive\"', 'code': '23505'}")
    import openpyxl as _x
    import tempfile, os
    fd, p = tempfile.mkstemp(suffix=".xlsx"); os.close(fd)
    wb = _x.Workbook(); ws = wb.active
    ws.append(["품번", "상품명", "색상", "사이즈", "도매가", "판매가"])
    ws.append(("5015", "이미있음", "블랙", "F", 1000, 2000))
    wb.save(p)
    out = ingest_excel(DupRepo(), "w1", p)
    os.unlink(p)
    assert out["errors"][0]["reason"] == "이미 등록된 품번입니다 (중복 — 건너뜀)"
    assert "23505" not in out["errors"][0]["reason"]      # raw DB 정보 미노출


def test_list_unmatched_returns_only_unmatched():
    repo = FakeUploadRepo()
    repo.create_upload_job({"wholesaler_id": "w1"})
    repo.insert_images([
        {"wholesaler_id": "w1", "storage_path": "a", "original_filename": "a.jpg",
         "product_id": None, "match_status": "unmatched"},
        {"wholesaler_id": "w1", "storage_path": "b", "original_filename": "b.jpg",
         "product_id": "p1", "match_status": "matched"},
    ])
    out = list_unmatched(repo, "job-1", caller_wid="w1")
    assert len(out) == 1 and out[0]["original_filename"] == "a.jpg"
