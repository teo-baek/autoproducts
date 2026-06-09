"""GET /products (목록·상세·엑셀) — 도매 본인 스코프 + 관리뷰 가격 셰이핑."""
from openpyxl import load_workbook
from io import BytesIO

from fastapi.testclient import TestClient
from app.main import app
from app.core.auth import get_current_user
from app.schemas.auth import CurrentUser
import app.routers.products as prod_mod


def _row(pid, wid, **over):
    base = {
        "id": pid, "wholesaler_id": wid, "platform_code": f"EZM-{pid}",
        "source_p_number": f"SRC-{pid}", "item_name": f"상품{pid}", "category": "의류",
        "fabric_composition": "Silk 100%", "status": "active",
        "product_skus": [
            {"id": f"s-{pid}", "color": "BLACK", "size": "FREE",
             "wholesale_price": 18000, "retail_price": 29000, "stock": 5, "deleted_at": None},
        ],
        "product_images": [],
    }
    base.update(over)
    return base


class FakeRepo:
    store = {}

    def __init__(self, owner_wid=None):
        self.owner_wid = owner_wid

    def list_products(self, *, limit, offset, category=None, search=None, status=None):
        rows = [r for r in FakeRepo.store.values() if r["wholesaler_id"] == self.owner_wid]
        if status:
            rows = [r for r in rows if r.get("status") == status]
        if category:
            rows = [r for r in rows if r.get("category") == category]
        if search:
            rows = [r for r in rows if search in r["item_name"] or search in r["source_p_number"]]
        total = len(rows)
        return rows[offset:offset + limit], total

    def get_product(self, pid):
        r = FakeRepo.store.get(pid)
        if not r or r["wholesaler_id"] != self.owner_wid:
            return None
        return r

    def replace_skus(self, pid, skus, updated_by=None):
        FakeRepo.store[pid]["product_skus"] = [
            {**s, "id": f"new-{i}", "deleted_at": None} for i, s in enumerate(skus)
        ]


def _patch(monkeypatch, rows):
    FakeRepo.store = {r["id"]: r for r in rows}
    monkeypatch.setattr(prod_mod, "SupabaseProductRepo", FakeRepo)


def _wholesaler(wid):
    return lambda: CurrentUser(id=f"u-{wid}", role="wholesaler", status="approved", wholesaler_id=wid)


def test_list_scoped_to_owner_with_management_price(monkeypatch):
    _patch(monkeypatch, [_row("1", "w1"), _row("2", "w1"), _row("9", "w9")])
    app.dependency_overrides[get_current_user] = _wholesaler("w1")
    try:
        r = TestClient(app).get("/products")
        assert r.status_code == 200
        body = r.json()
        assert body["total"] == 2                       # 타 업체(w9) 제외
        sku = body["items"][0]["skus"][0]
        assert sku["wholesale_price"] == 18000          # 본인 → 관리뷰(도매가+판매가)
        assert sku["retail_price"] == 29000
        assert "price" not in sku                        # 단일가 셰이핑 아님
    finally:
        app.dependency_overrides.clear()


def test_list_category_filter(monkeypatch):
    _patch(monkeypatch, [_row("1", "w1", category="의류"), _row("2", "w1", category="잡화")])
    app.dependency_overrides[get_current_user] = _wholesaler("w1")
    try:
        r = TestClient(app).get("/products?category=잡화")
        assert r.json()["total"] == 1
        assert r.json()["items"][0]["category"] == "잡화"
    finally:
        app.dependency_overrides.clear()


def test_detail_foreign_owner_404(monkeypatch):
    _patch(monkeypatch, [_row("9", "w9")])
    app.dependency_overrides[get_current_user] = _wholesaler("w1")
    try:
        assert TestClient(app).get("/products/9").status_code == 404
    finally:
        app.dependency_overrides.clear()


def test_detail_owner_ok(monkeypatch):
    _patch(monkeypatch, [_row("1", "w1")])
    app.dependency_overrides[get_current_user] = _wholesaler("w1")
    try:
        r = TestClient(app).get("/products/1")
        assert r.status_code == 200
        assert r.json()["platform_code"] == "EZM-1"
    finally:
        app.dependency_overrides.clear()


def test_replace_skus_owner_ok(monkeypatch):
    _patch(monkeypatch, [_row("1", "w1")])
    app.dependency_overrides[get_current_user] = _wholesaler("w1")
    try:
        r = TestClient(app).put("/products/1/skus", json={"skus": [
            {"color": "RED", "size": "M", "wholesale_price": 20000, "retail_price": 30000, "stock": 3}]})
        assert r.status_code == 200
        skus = r.json()["skus"]
        assert len(skus) == 1 and skus[0]["color"] == "RED"
        assert skus[0]["wholesale_price"] == 20000
    finally:
        app.dependency_overrides.clear()


def test_replace_skus_foreign_404(monkeypatch):
    _patch(monkeypatch, [_row("9", "w9")])
    app.dependency_overrides[get_current_user] = _wholesaler("w1")
    try:
        r = TestClient(app).put("/products/9/skus", json={"skus": [
            {"color": "RED", "size": "M", "wholesale_price": 1, "retail_price": None, "stock": 0}]})
        assert r.status_code == 404
    finally:
        app.dependency_overrides.clear()


def test_export_xlsx_owner(monkeypatch):
    _patch(monkeypatch, [_row("1", "w1")])
    app.dependency_overrides[get_current_user] = _wholesaler("w1")
    try:
        r = TestClient(app).get("/products/export.xlsx")
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers["content-type"]
        wb = load_workbook(BytesIO(r.content))
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "사진"   # 스타일 A~L: A열=사진
        assert [c.value for c in ws[1]][-1] == "QR 이미지"  # 최우측 열 = QR 이미지(PNG)
        assert ws.cell(row=2, column=2).value == "SRC-1"  # B 품번
        assert ws.cell(row=2, column=7).value == 18000     # G 도매가
        assert ws.cell(row=2, column=8).value == 29000     # H 판매가(관리뷰=둘 다)
        assert ws.cell(row=2, column=10).value == "18_29"  # J P CODE (관리뷰=둘 다 보임)
        assert ws.cell(row=2, column=1).value == "사진 없음"  # 이미지 없는 행
        # K(QR 링크) = URL 텍스트(품번 포함)
        qr = ws.cell(row=2, column=11).value
        assert isinstance(qr, str) and "EZM-1" in qr
        # L(QR 이미지) = QR PNG 임베드 → 사진 없는 행이라도 이미지 1개(QR)
        assert len(ws._images) == 1
    finally:
        app.dependency_overrides.clear()
