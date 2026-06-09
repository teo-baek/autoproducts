import io

import openpyxl
from PIL import Image as PILImage

from app.services.excel_export import (
    build_catalog_xlsx, build_render_xlsx, RENDER_HEADERS, cell_image_path, cell_image_bytes,
)
from app.services.pricing import visible_price_columns


def test_export_has_qr_column_last(tmp_path):
    items = [{"platform_code": "EZM-000001", "item_name": "린넨셔츠", "price": 12000}]
    out = tmp_path / "out.xlsx"
    build_catalog_xlsx(items, str(out), base_url="https://x")
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert headers[-1] == "QR"            # 최우측 열이 QR (FR-3.2)
    assert ws.cell(row=2, column=1).value == "EZM-000001"
    assert len(ws._images) == 1           # QR 이미지 1개 삽입됨


# ── 스타일 렌더 엑셀(사진·QR 박은 A~J) ──────────────────────────────────────────
def _render_one(skus, image_bytes=None):
    rows = [{
        "source_p_number": "A-001", "item_name": "린넨셔츠", "fabric_composition": "Linen 100%",
        "platform_code": "EZM-000001", "image_bytes": image_bytes, "skus": skus,
    }]
    return openpyxl.load_workbook(io.BytesIO(build_render_xlsx(rows, base_url="https://x"))).active


def test_render_layout_headers_and_qr_last():
    ws = _render_one([{"color": "화이트", "size": "S", "stock": 10,
                       "wholesale_price": 12000, "retail_price": 29000}])
    assert [c.value for c in ws[1]] == RENDER_HEADERS           # A~L 12열 (P CODE·QR 링크·QR 이미지 포함)
    assert [c.value for c in ws[1]][-1] == "QR 이미지"          # 최우측 = QR 이미지(PNG)
    assert ws.cell(2, 2).value == "A-001"                       # B 품번
    assert ws.cell(2, 3).value == "린넨셔츠"                     # C 상품명
    assert ws.cell(2, 6).value == "Linen 100%"                 # F 혼용률
    assert ws.cell(2, 7).value == 12000                        # G 도매가
    assert ws.cell(2, 8).value == 29000                        # H 판매가
    assert ws.cell(2, 9).value == 10                           # I 재고
    assert ws.cell(2, 10).value == "12_29"                     # J P CODE = 도매//1000_소매//1000
    # K(QR 링크) = URL 텍스트(품번 포함)
    qr_link = ws.cell(2, 11).value
    assert isinstance(qr_link, str) and "EZM-000001" in qr_link
    assert ws.cell(2, 1).value == "사진 없음"                   # A 이미지 없음
    assert len(ws._images) == 1                                # L QR 이미지 1개(사진 없음)


def test_render_repeats_product_cells_per_sku():
    ws = _render_one([
        {"color": "화이트", "size": "S", "stock": 10, "wholesale_price": 12000, "retail_price": 29000},
        {"color": "화이트", "size": "M", "stock": 5, "wholesale_price": 12000, "retail_price": None},
    ])
    # SKU 단위 값은 행마다
    assert ws.cell(2, 5).value == "S" and ws.cell(3, 5).value == "M"
    assert ws.cell(2, 9).value == 10 and ws.cell(3, 9).value == 5
    assert ws.cell(3, 8).value in (None, "")                   # 둘째 SKU 판매가 None → 공백
    # P CODE: 둘 다 있는 행만 생성, 한쪽 None 이면 빈칸(가격 유출 방지)
    assert ws.cell(2, 10).value == "12_29"
    assert ws.cell(3, 10).value in (None, "")
    # 상품 공통(품번·상품명·혼용률)은 병합 없이 매 행 반복
    assert ws.cell(2, 2).value == "A-001" and ws.cell(3, 2).value == "A-001"
    assert ws.cell(2, 3).value == "린넨셔츠" and ws.cell(3, 3).value == "린넨셔츠"
    assert ws.cell(2, 6).value == "Linen 100%" and ws.cell(3, 6).value == "Linen 100%"
    assert len(ws.merged_cells.ranges) == 0                    # 병합 없음
    assert len(ws._images) == 2                                # L QR 이미지 행마다(사진 없음) → 2개


def test_render_embeds_image_when_present():
    buf = io.BytesIO(); PILImage.new("RGB", (8, 8), "red").save(buf, format="PNG")
    ws = _render_one([{"color": "c", "size": "s", "stock": 1,
                       "wholesale_price": 1, "retail_price": 1}], image_bytes=buf.getvalue())
    assert ws.cell(2, 1).value is None                         # '사진 없음' 아님
    assert len(ws._images) == 2                                # 사진 + QR


# ── visible_price_columns: 역할별 2칸(도매가/판매가) 정규화 ─────────────────────
_SKU = {"wholesale_price": 10000, "retail_price": 25000, "product_org": "org-1"}


def test_price_columns_admin_and_owner_show_both():
    both = {"wholesale_price": 10000, "retail_price": 25000}
    assert visible_price_columns("admin", None, _SKU) == both
    assert visible_price_columns("wholesaler", None, _SKU, viewer_org="org-1") == both


def test_price_columns_independent_seller_wholesale_only():
    assert visible_price_columns("retail_seller", "independent", _SKU) == {
        "wholesale_price": 10000, "retail_price": None}


def test_price_columns_agency_retail_only():
    assert visible_price_columns("agency", None, _SKU) == {
        "wholesale_price": None, "retail_price": 25000}


def test_price_columns_hidden_both_none():
    assert visible_price_columns("retail_seller", "agency_affiliated", _SKU) == {
        "wholesale_price": None, "retail_price": None}
    assert visible_price_columns("retail_seller", "independent", _SKU,
                                 price_visibility="none") == {
        "wholesale_price": None, "retail_price": None}


# ── 엑셀 export 이미지 경로/캐시 (audit Critical 수정) ──────────────────────────
def test_cell_image_path_prefers_thumbnail():
    assert cell_image_path({"storage_path": "w/orig.jpg", "thumbnail_path": "thumbs/w/orig.jpg"}) == "thumbs/w/orig.jpg"
    assert cell_image_path({"storage_path": "w/orig.jpg", "thumbnail_path": None}) == "w/orig.jpg"  # 썸네일 없으면 원본 fallback
    assert cell_image_path({"storage_path": "w/orig.jpg"}) == "w/orig.jpg"
    assert cell_image_path(None) is None


class _FakeStorage:
    def __init__(self, data):
        self.data = data
        self.calls = 0
    def from_(self, _bucket):
        return self
    def download(self, _path):
        self.calls += 1
        return self.data


class _FakeSB:
    def __init__(self, data):
        self.storage = _FakeStorage(data)


def test_cell_image_bytes_caches_by_path():
    buf = io.BytesIO(); PILImage.new("RGB", (20, 20), "red").save(buf, format="PNG")
    sb = _FakeSB(buf.getvalue())
    path = "w/cache-test-unique-987.jpg"   # 전역 캐시 충돌 방지용 고유 경로
    a = cell_image_bytes(sb, path)
    b = cell_image_bytes(sb, path)
    assert a and b and a == b
    assert sb.storage.calls == 1           # 두 번째는 TTL 캐시 → 재다운로드 안 함


def test_cell_image_bytes_none_path():
    assert cell_image_bytes(_FakeSB(b""), None) is None
