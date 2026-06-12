import io
import openpyxl
from fastapi.testclient import TestClient
from app.main import app
from app.core.auth import get_current_user
from app.schemas.auth import CurrentUser
from app.services.excel_export import catalog_xlsx_bytes
import app.routers.catalog as catalog_mod


def test_catalog_xlsx_bytes_has_qr_header():
    data = catalog_xlsx_bytes(
        [{"platform_code": "EZM-000001", "item_name": "셔츠", "price": 12000}],
        base_url="https://x",
    )
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active
    assert [c.value for c in ws[1]][-1] == "QR"        # 최우측 열 QR (FR-3.2)
    assert ws.cell(row=2, column=1).value == "EZM-000001"


def test_export_route_returns_xlsx_with_role_shaped_price(monkeypatch):
    rows = [{
        "platform_code": "EZM-000001", "item_name": "셔츠", "source_p_number": "SRC-1",
        "fabric_composition": "Cotton 100%", "wholesaler_id": "org-9",
        "representative_image_url": None, "product_images": [],
        "product_skus": [{"color": "화이트", "size": "F",
                          "wholesale_price": 12000, "retail_price": 29000, "stock": 3}],
    }]
    monkeypatch.setattr(catalog_mod, "_query_catalog_export_rows", lambda sb, limit, wholesaler_ids=None: rows)
    monkeypatch.setattr(catalog_mod, "get_supabase", lambda: object())
    app.dependency_overrides[get_current_user] = lambda: CurrentUser(
        id="u", role="retail_seller", status="approved", seller_type="independent")
    try:
        res = TestClient(app).get("/catalog/export.xlsx")
        assert res.status_code == 200
        assert "spreadsheet" in res.headers["content-type"]
        assert "attachment" in res.headers["content-disposition"]
        wb = openpyxl.load_workbook(io.BytesIO(res.content))
        ws = wb.active
        assert [c.value for c in ws[1]][-1] == "QR 이미지"    # 최우측 열 = QR 이미지(PNG)
        # independent 셀러 → 도매가(G)만 노출, 판매가(H)는 공백 (FR-5.2)
        assert ws.cell(row=2, column=7).value == 12000        # G 도매가
        assert ws.cell(row=2, column=8).value in (None, "")   # H 판매가 미노출
        # P CODE(J) = 가격코드 '도매앞2_판매앞2' — 라이브셀러 작업용(가격 하나라도 보이면 노출, 2026-06-07)
        assert ws.cell(row=2, column=10).value == "12_29"
        # K(QR 링크) = 링크 URL 텍스트(품번 포함)
        qr = ws.cell(row=2, column=11).value
        assert isinstance(qr, str) and "EZM-000001" in qr
        # L(QR 이미지) = QR PNG 임베드(사진 없는 행이라도 QR 이미지 1개)
        assert len(ws._images) == 1
    finally:
        app.dependency_overrides.clear()


def test_export_pcode_blank_for_price_hidden_seller(monkeypatch):
    """가격 완전 미노출(agency_affiliated=none) 셀러 → 도매·판매·P CODE 모두 빈칸(코드 유출 방지)."""
    rows = [{
        "platform_code": "EZM-000001", "item_name": "셔츠", "source_p_number": "SRC-1",
        "fabric_composition": None, "wholesaler_id": "org-9",
        "representative_image_url": None, "product_images": [],
        "product_skus": [{"color": "화이트", "size": "F",
                          "wholesale_price": 12000, "retail_price": 29000, "stock": 3}],
    }]
    monkeypatch.setattr(catalog_mod, "_query_catalog_export_rows", lambda sb, limit, wholesaler_ids=None: rows)
    monkeypatch.setattr(catalog_mod, "get_supabase", lambda: object())
    app.dependency_overrides[get_current_user] = lambda: CurrentUser(
        id="u", role="retail_seller", status="approved", seller_type="agency_affiliated")
    try:
        res = TestClient(app).get("/catalog/export.xlsx")
        assert res.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(res.content))
        ws = wb.active
        assert ws.cell(row=2, column=7).value in (None, "")    # 도매가 미노출
        assert ws.cell(row=2, column=8).value in (None, "")    # 판매가 미노출
        assert ws.cell(row=2, column=10).value in (None, "")   # P CODE 빈칸
    finally:
        app.dependency_overrides.clear()


def test_export_route_blocks_unapproved(monkeypatch):
    app.dependency_overrides[get_current_user] = lambda: CurrentUser(
        id="u", role="retail_seller", status="pending", seller_type="independent")
    try:
        res = TestClient(app).get("/catalog/export.xlsx")
        assert res.status_code == 403   # 미승인 → 차단 (FR-5.1)
    finally:
        app.dependency_overrides.clear()


def test_catalog_list_blocks_wholesaler():
    """도매사↔도매사 격리: wholesaler 는 셀러용 카탈로그를 볼 수 없다 — 같은 테넌트 타 도매사 상품·이미지 노출 차단."""
    app.dependency_overrides[get_current_user] = lambda: CurrentUser(
        id="w", role="wholesaler", status="approved", wholesaler_id="org-A")
    try:
        res = TestClient(app).get("/catalog")
        assert res.status_code == 403   # 셀러 전용 — 도매사 차단
    finally:
        app.dependency_overrides.clear()


def test_export_route_blocks_wholesaler():
    """엑셀 출력도 동일 격리 — wholesaler 가 호출하면 타 도매사 이미지 바이트가 새어나가므로 차단."""
    app.dependency_overrides[get_current_user] = lambda: CurrentUser(
        id="w", role="wholesaler", status="approved", wholesaler_id="org-A")
    try:
        res = TestClient(app).get("/catalog/export.xlsx")
        assert res.status_code == 403
    finally:
        app.dependency_overrides.clear()
