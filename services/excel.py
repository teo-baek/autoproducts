from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as OpenpyxlImage
import io

class ExcelGenerator:
    def __init__(self):
        pass

    def generate_excel(self, parsed_rows_data, image_results, qr_results=None):
        """파싱된 데이터와 이미지 버퍼 리스트를 매핑하여 엑셀 파일을 생성하고 BytesIO 반환"""
        wb = Workbook()
        ws = wb.active
        ws.title = "방송제품목록"

        headers = ["사진", "품번", "상품명", "색상", "상세사이즈", "혼용률", "도매가", "판매가", "재고", "P CODE", "모바일 스캔 (QR)"]
        ws.append(headers)

        for idx, data_row in enumerate(parsed_rows_data):
            ws.append([
                "", data_row["p_number"], data_row["item_name"], data_row["color"],
                data_row["size"], data_row["mix_ratio"], data_row["wholesale"],
                data_row["retail"], data_row["stock"], data_row["p_code"]
            ])
            
            current_row_idx = idx + 2
            
            # 메인 상품 이미지 삽입 (A열)
            img_res = image_results[idx]
            if isinstance(img_res, io.BytesIO):
                try:
                    xl_img = OpenpyxlImage(img_res)
                    ws.add_image(xl_img, f"A{current_row_idx}")
                except:
                    ws.cell(row=current_row_idx, column=1, value="이미지 오류")
            else:
                ws.cell(row=current_row_idx, column=1, value="사진 없음")
                
            # QR 코드 이미지 삽입 (K열)
            if qr_results and len(qr_results) > idx:
                qr_res = qr_results[idx]
                if isinstance(qr_res, io.BytesIO):
                    try:
                        qr_xl_img = OpenpyxlImage(qr_res)
                        ws.add_image(qr_xl_img, f"K{current_row_idx}")
                    except:
                        ws.cell(row=current_row_idx, column=11, value="QR 오류")

        # --- 디자이너 무드 스타일링 공정 ---
        font_main = Font(name="맑은 고딕", size=10)
        font_header = Font(name="맑은 고딕", size=11, bold=True)
        
        fill_header = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

        thin_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )

        ws.row_dimensions[1].height = 28
        for cell in ws[1]:
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = thin_border

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=11), start=2):
            ws.row_dimensions[row_idx].height = 160
            for col_idx, cell in enumerate(row, start=1):
                cell.font = font_main
                cell.border = thin_border
                cell.alignment = align_center

                if col_idx in [3, 5, 6]:
                    cell.alignment = align_left

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['K'].width = 25  # QR코드 열 너비 확보
        for col in ws.iter_cols(min_col=2, max_col=10):
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 4, 13)
        
        excel_data = io.BytesIO()
        wb.save(excel_data)
        excel_data.seek(0)
        
        return excel_data
