import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as OpenpyxlImage
import io
import re
import urllib.request
import json
from PIL import Image as PILImage, ImageOps
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor
import qrcode
import urllib.parse

# 스트림릿 페이지 설정
st.set_page_config(page_title="제품 관리 자동화", layout="wide")

def get_q_param(key):
    val = st.query_params.get(key, "")
    if isinstance(val, list):
        return val[0] if val else ""
    return val

# --- 모바일 뷰어 라우팅 (QR 스캔 시) ---
if get_q_param("view") == "qr":
    st.markdown("""
        <style>
        #MainMenu {visibility: hidden;}
        footer {visibility: hidden;}
        header {visibility: hidden;}
        .block-container { padding-top: 1rem; padding-bottom: 1rem; max-width: 100%; }
        </style>
    """, unsafe_allow_html=True)

    f_id = get_q_param("f_id")
    folder = get_q_param("folder")
    p_num = get_q_param("p")
    stk = get_q_param("stk")
    nm = get_q_param("nm")
    
    st.markdown(f"<h1 style='text-align: center; margin-bottom: 5px; font-weight: 800;'>{nm}</h1>", unsafe_allow_html=True)
    st.markdown(f"<h3 style='text-align: center; color: #666; margin-top: 0;'>품번: {p_num}</h3>", unsafe_allow_html=True)
    
    img_data = None
    
    # 1. file_id가 있을 경우 (uc 원본 다운로드 폐기, 초고속 썸네일 엔진 단일 적용)
    if f_id:
        try:
            req = urllib.request.Request(f"https://drive.google.com/thumbnail?id={f_id}&sz=w600", headers={'User-Agent': 'Mozilla/5.0'})
            with urllib.request.urlopen(req, timeout=3) as response:
                img_data = response.read()
        except:
            pass
                
    # 2. file_id가 없고 폴더 기반 우회 스캔이 필요한 경우 (동일하게 초고속 썸네일 엔진 적용)
    if not img_data and folder and p_num:
        for ext in ['jpg', 'jpeg', 'png', 'JPG', 'JPEG', 'PNG']:
            fallback_url = f"https://drive.google.com/thumbnail?authuser=0&sz=w600&id={folder}&filename={p_num}.{ext}"
            try:
                req = urllib.request.Request(fallback_url, headers={'User-Agent': 'Mozilla/5.0'})
                with urllib.request.urlopen(req, timeout=2) as response:
                    img_data = response.read()
                    if img_data:
                        break
            except:
                continue

    # 3. 이미지 렌더링 및 서버 사이드 정면 고정 (PIL EXIF Transpose)
    if img_data:
        try:
            pil_img = PILImage.open(io.BytesIO(img_data))
            pil_img = ImageOps.exif_transpose(pil_img)
            st.image(pil_img, use_container_width=True)
        except Exception:
            # 안전 장치: 처리 중 에러 발생 시 원본 썸네일 데이터로 출력
            st.image(img_data, use_container_width=True)
    else:
        st.info("해당 제품의 이미지를 불러올 수 없습니다.")
        
    st.markdown(f"""
        <div style='background: linear-gradient(135deg, #FF416C 0%, #FF4B2B 100%); padding: 20px; border-radius: 15px; text-align: center; margin-top: 20px; box-shadow: 0 4px 15px rgba(255, 75, 43, 0.4);'>
            <h4 style='margin: 0; color: rgba(255,255,255,0.9); font-size: 16px;'>현재고</h4>
            <h2 style='margin: 5px 0 0 0; color: white; font-size: 40px; font-weight: 900;'>{stk} <span style='font-size: 20px;'>개</span></h2>
        </div>
    """, unsafe_allow_html=True)
    
    st.stop() # 모바일 뷰어 모드일 경우 여기서 앱 실행 종료

st.title("제품 관리 엑셀 자동화")
st.markdown("포스기에서 다운받은 엑셀 파일(raw)과 구글 드라이브 폴더 주소를 입력하면, 셀 내부에 이미지가 삽입된 엑셀을 생성합니다.")

# 1. 파일 및 정보 입력 UI
st.subheader("📋 1. 데이터 입력")
uploaded_file = st.file_uploader("매장 포스기 제품 엑셀 파일을 업로드하세요. (XLSX, XLS, CSV 지원)", type=["csv", "xlsx", "xls"])
folder_url = st.text_input("제품 사진이 업로드되어 있는 구글 드라이브 폴더 주소(URL)를 입력하세요.")

# 스트림릿 클라우드에 배포된 고정 퍼블릭 주소 하드코딩
app_url = "https://autoappucts-mesgtt2ne6wscqus6ggnrq.streamlit.app"

def extract_folder_id(url):
    """구글 드라이브 URL에서 폴더 고유 ID를 추출하는 함수"""
    match = re.search(r"folders/([a-zA-Z0-9-_]+)", url)
    if match:
        return match.group(1)
    return url.strip()

def clean_and_parse_price(price_val):
    """가격 데이터에서 쉼표나 문자를 제거하고 정수로 변환하는 함수"""
    if pd.isna(price_val):
        return 0
    try:
        price_str = re.sub(r'[^\d]', '', str(price_val))
        return int(price_str) if price_str else 0
    except:
        return 0

def get_column_value_by_synonyms(row, df_columns, synonyms, default_val=""):
    """매장별로 다른 포스기 열 이름을 동의어 풀을 순회하며 감지하는 함수"""
    for synonym in synonyms:
        matched_col = [col for col in df_columns if str(col).strip() == synonym]
        if matched_col:
            val = row[matched_col[0]]
            return str(val).strip() if pd.notna(val) else default_val
    return default_val

def get_google_drive_file_list(folder_id):
    """Google Apps Script 웹 앱을 통해 구글 드라이브 폴더 내 파일명과 ID 목록을 안정적으로 가져오는 함수"""
    file_map = {}
    try:
        gas_url = f"https://script.google.com/macros/s/AKfycbwkKOJJeZX75jmWXP-gaXw__cyec6tXxKYQ8cxp8Ou5emWvXhN6KedCH0j3mkZPcl3L1w/exec?folderId={folder_id}"
        req = urllib.request.Request(gas_url)
        with urllib.request.urlopen(req, timeout=15) as response:
            data = json.loads(response.read().decode('utf-8'))
            for f_name, f_id in data.items():
                # 확장자 제거 (예: product.jpg -> product)
                clean_name = re.sub(r'\.[a-zA-Z0-9]+$', '', f_name)
                # 포스기 품번 형태에 맞춰 .0 오차 제거 및 공백 제거
                clean_key = re.sub(r'\.0$', '', clean_name.strip())
                file_map[clean_key] = f_id
    except:
        pass
    return file_map

def download_thumbnail_image(file_id):
    """구글 드라이브 파일 ID를 이용해 썸네일 또는 원본 이미지를 다운로드하는 함수"""
    if not file_id:
        return None
        
    # 1차 시도: UC (다이렉트) 엔드포인트
    uc_url = f"https://drive.google.com/uc?id={file_id}"
    try:
        req = urllib.request.Request(uc_url, headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'})
        with urllib.request.urlopen(req, timeout=5) as response:
            return response.read()
    except:
        pass
        
    # 2차 시도: 썸네일 엔드포인트
    thumb_url = f"https://drive.google.com/thumbnail?id={file_id}&sz=w300"
    try:
        req = urllib.request.Request(thumb_url, headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'})
        with urllib.request.urlopen(req, timeout=4) as response:
            return response.read()
    except:
        return None

# NameError 및 타겟 오차를 차단하고 스트리밍 채널로 고도화한 단일 스레드 타겟 함수
def process_single_row_image(p_num, file_list_map, folder_id):
    """5015.0 실수형 변환 오차를 정규식으로 완벽 분리 후 이미지 매칭 및 다운로드"""
    if not p_num:
        return {"image": "NONE", "file_id": ""}
    
    # 1단계: 판다스가 실수로 붙인 소수점 .0을 완벽히 도려내 정형화 (NameError 버그 완치)
    clean_p_num = re.sub(r'\.0$', '', str(p_num).strip())
    
    # 2단계: 스캔된 파일 리스트에서 ID 획득 시도
    file_id = file_list_map.get(clean_p_num)
    
    img_data = None
    if file_id:
        img_data = download_thumbnail_image(file_id)
        
    # 3단계: 파일 ID 스캔이 실패했거나 구글 웹 토큰이 막혔을 때 작동하는 '주소 스캔' 레이어
    if not img_data:
        img_extensions = ['jpg', 'jpeg', 'png', 'JPG', 'JPEG', 'PNG']
        for ext in img_extensions:
            # uc?export가 아닌 보안 차단이 없는 썸네일 다이렉트 뷰어 구조를 조합하여 구글 서버를 우회 관통합니다.
            fallback_url = f"https://drive.google.com/thumbnail?authuser=0&sz=w300&id={folder_id}&filename={clean_p_num}.{ext}"
            try:
                req = urllib.request.Request(fallback_url, headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'})
                with urllib.request.urlopen(req, timeout=2) as response:
                    img_data = response.read()
                    if img_data:
                        break
            except:
                continue
                
    if img_data:
        try:
            pil_img = PILImage.open(io.BytesIO(img_data))
            pil_img = ImageOps.exif_transpose(pil_img)  # EXIF 방향 메타데이터를 실제 픽셀에 물리적으로 적용
            pil_img = pil_img.convert("RGB")
            pil_img.thumbnail((90, 110))
            
            img_buffer = io.BytesIO()
            pil_img.save(img_buffer, format="JPEG")
            img_buffer.seek(0)
            return {"image": img_buffer, "file_id": file_id if file_id else ""}
        except:
            return {"image": "ERROR", "file_id": file_id if file_id else ""}
            
    return {"image": "NONE", "file_id": file_id if file_id else ""}

# 2. 변환 프로세스 시작
if st.button("파일 변환"):
    if uploaded_file is not None and folder_url:
        try:
            folder_id = extract_folder_id(folder_url)

            # 포스기 구형/신형 엑셀 파싱
            if uploaded_file.name.endswith('.csv'):
                df_raw = pd.read_csv(uploaded_file, encoding='utf-8')
            elif uploaded_file.name.endswith('.xls'):
                df_raw = pd.read_excel(uploaded_file, engine='xlrd')
            else:
                df_raw = pd.read_excel(uploaded_file)
                
            st.info("📊 구글 드라이브 폴더의 파일 보안 토큰 구조를 실시간 스캔하는 중입니다...")
            
            # 구글 폴더 내 파일명과 고유 ID 일괄 선행 추출
            file_list_map = get_google_drive_file_list(folder_id)
            
            cols = df_raw.columns
            synonyms_pool = {
                "p_number": ['품번', '상품코드', '품목코드', '모델명'],
                "item_name": ['상품명', '품목명', '물품명', '제품명'],
                "color": ['색상', '컬러', '색상명'],
                "size": ['상세사이즈', '사이즈', '규격'],
                "mix_ratio": ['혼용률', '혼방률', '소재'],
                "wholesale": ['도매가', '도매단가', '입고가', '공급가'],
                "retail": ['소매가', '판매가', '소비자가', '매장판매가'],
                "stock": ['재고정상', '재고', '현재고', '매장량', '수량']
            }

            parsed_rows_data = []

            # 1차 데이터 파싱 및 정형화
            for idx, row in df_raw.iterrows():
                p_num_raw = get_column_value_by_synonyms(row, cols, synonyms_pool["p_number"])
                p_num = re.sub(r'\.0$', '', str(p_num_raw).strip())
                
                name = get_column_value_by_synonyms(row, cols, synonyms_pool["item_name"])
                col_val = get_column_value_by_synonyms(row, cols, synonyms_pool["color"])
                sz_val = get_column_value_by_synonyms(row, cols, synonyms_pool["size"], default_val="F")
                mix_val = get_column_value_by_synonyms(row, cols, synonyms_pool["mix_ratio"])
                
                wholesale_raw = get_column_value_by_synonyms(row, cols, synonyms_pool["wholesale"], default_val="0")
                retail_raw = get_column_value_by_synonyms(row, cols, synonyms_pool["retail"], default_val="0")
                
                wholesale_num = clean_and_parse_price(wholesale_raw)
                retail_num = clean_and_parse_price(retail_raw)
                
                wholesale_text = f"{wholesale_num:,}" if wholesale_num > 0 else "0"
                retail_text = f"{retail_num:,}" if retail_num > 0 else "0"
                
                stk_val = get_column_value_by_synonyms(row, cols, synonyms_pool["stock"], default_val="0")
                stk_val = re.sub(r'\.0$', '', stk_val)
                
                # P_CODE 연산 후 최종 값 주입 방식 유지
                p_code_num1 = int(wholesale_num / 1000)
                p_code_num2 = int(retail_num / 1000)
                p_code_text = f"{p_code_num1}_{p_code_num2}"
                
                parsed_rows_data.append({
                    "p_number": p_num, "item_name": name, "color": col_val, "size": sz_val,
                    "mix_ratio": mix_val, "wholesale": wholesale_text, "retail": retail_text,
                    "stock": stk_val, "p_code": p_code_text
                })

            st.info(f"이미지 {len(parsed_rows_data)}개를 삽입합니다...")
            
            # 이미지 매칭 및 예외 백업을 종합적으로 처리하는 마스터 함수를 비동기 병렬 구조로 타겟팅 변경
            image_results = []
            with ThreadPoolExecutor(max_workers=24) as executor:
                futures = [
                    executor.submit(process_single_row_image, data["p_number"], file_list_map, folder_id) 
                    for data in parsed_rows_data
                ]
                
                progress_bar = st.progress(0)
                for f_idx, future in enumerate(futures):
                    image_results.append(future.result())
                    progress_bar.progress((f_idx + 1) / len(futures))

            # openpyxl 워크북 마스터 렌더링 시작
            wb = Workbook()
            ws = wb.active
            ws.title = "방송제품목록"

            headers = ["사진", "품번", "상품명", "색상", "상세사이즈", "혼용률", "도매가", "판매가", "재고", "P CODE", "QR 링크"]
            ws.append(headers)

            for idx, data_row in enumerate(parsed_rows_data):
                ws.append([
                    "", data_row["p_number"], data_row["item_name"], data_row["color"],
                    data_row["size"], data_row["mix_ratio"], data_row["wholesale"],
                    data_row["retail"], data_row["stock"], data_row["p_code"], ""
                ])
                
                current_row_idx = idx + 2
                img_res = image_results[idx]
                
                img_buffer = img_res["image"]
                f_id = img_res["file_id"]
                
                # 제품 사진 삽입
                if isinstance(img_buffer, io.BytesIO):
                    try:
                        xl_img = OpenpyxlImage(img_buffer)
                        ws.add_image(xl_img, f"A{current_row_idx}")
                    except:
                        ws.cell(row=current_row_idx, column=1, value="이미지 오류")
                else:
                    ws.cell(row=current_row_idx, column=1, value="사진 없음")
                    
                # QR 링크 생성 및 삽입
                qr_url = f"{app_url.rstrip('/')}/?view=qr&f_id={str(f_id)}&folder={str(folder_id)}&p={urllib.parse.quote(str(data_row['p_number']))}&stk={urllib.parse.quote(str(data_row['stock']))}&nm={urllib.parse.quote(str(data_row['item_name']))}"
                try:
                    qr = qrcode.QRCode(box_size=3, border=1)
                    qr.add_data(qr_url)
                    qr.make(fit=True)
                    qr_img = qr.make_image(fill_color="black", back_color="white")
                    
                    qr_buffer = io.BytesIO()
                    qr_img.save(qr_buffer, format="PNG")
                    qr_buffer.seek(0)
                    
                    xl_qr = OpenpyxlImage(qr_buffer)
                    # 엑셀 셀 안에 쏙 들어가도록 이미지 사이즈 강제 조정 (가로/세로 100px)
                    xl_qr.width = 100
                    xl_qr.height = 100
                    
                    ws.add_image(xl_qr, f"K{current_row_idx}")
                except Exception as e:
                    ws.cell(row=current_row_idx, column=11, value="QR 오류")

            # --- 디자이너 무드 스타일링 공정 ---
            font_main = Font(name="맑은 고딕", size=10)
            font_header = Font(name="맑은 고딕", size=11, bold=True)
            
            fill_header = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

            thin_border = Border(
                left=Side(style='thin', color='D3D3D3'),
                right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'),
                bottom=Side(style='thin', color='D3D3D3')
            )

            ws.row_dimensions[1].height = 28
            for cell in ws[1]:
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = thin_border

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=10), start=2):
                ws.row_dimensions[row_idx].height = 95
                for col_idx, cell in enumerate(row, start=1):
                    cell.font = font_main
                    cell.border = thin_border
                    cell.alignment = align_center

                    if col_idx in [3, 5, 6]:
                        cell.alignment = align_left

            ws.column_dimensions['A'].width = 16
            ws.column_dimensions['K'].width = 16
            for col in ws.iter_cols(min_col=2, max_col=10):
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = col[0].column_letter
                ws.column_dimensions[col_letter].width = max(max_len + 4, 13)
            
            excel_data = io.BytesIO()
            wb.save(excel_data)
            excel_data.seek(0)

            today_str = datetime.now().strftime("%Y-%m-%d")
            final_file_name = f"{today_str}_방송제품목록.xlsx"

            st.balloons()
            st.subheader("🎉 2. 세팅 완료 및 다운로드")
            st.download_button(
                label="🟢 변환 완료 및 세팅된 엑셀 파일 다운로드",
                data=excel_data,
                file_name=final_file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            st.error(f"변환 중 에러가 발생했습니다. 에러 내용: {e}")
    else:
        st.warning("포스기 제품 엑셀 파일과 구글 드라이브 폴더 주소를 누락 없이 입력해 주세요.")